#!/usr/bin/env python3
"""Build explanation_script.xlsx — the SA explanation stimulus pool as a workbook.

Same content as explanation_script.md, in a form that is easier to filter, sort and
diff against the HMI code. Korean strings are transcribed from:
  ../hmi-visual/src/App.jsx          SEQUENCES.roundabout / .aquaplaning (hero, sub)
  ../hmi-voice/src/data/drivePhases.js   C1_PHASES / C2_PHASES (speech)
English is the authored caption set used in the paper and the submission video.

    python3 build_explanation_xlsx.py     -> explanation_script.xlsx
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "explanation_script.xlsx"

# Status colours mirror the HMI's own AutopilotStatus chip (drivePhases.js)
GREEN, RED, ORANGE, YELLOW = "21C46A", "EF4444", "F97316", "EAB308"
INK, SEC, TER = "15181E", "5A636E", "9BA3AE"
FI_BAND = "FDEEE3"          # peach band on the FI back-inference rows
HEAD_BG = "F4F6F8"
RULE = "D8DCE1"

KO = "Malgun Gothic"        # renders Hangul on Windows and macOS Excel
EN = "Calibri"

thin = Side(style="thin", color=RULE)
BORDER = Border(bottom=thin)

STATUS_COLOR = {"Normal": GREEN, "Detect": RED, "Cause": ORANGE, "Resolve": YELLOW}

# ---------------------------------------------------------------- content
NORMAL_KO = "목적지까지 안전하게 주행 중입니다."
NORMAL_EN = "Driving safely toward the destination."
NONE = "— no explanation —"

C1 = [
    # phase, status, event, ZI-en, ZI-ko, ZO-en, ZO-ko, voice
    ("C1-1", "Normal", "drive_start, junction_arrive", NORMAL_EN, NORMAL_KO, NONE, "", "Short form: 정상 주행 중입니다"),
    ("C1-2", "Detect", "gap_attempt (attempt_n ≥ 2)",
     "Struggling to secure a gap to enter the junction.", "교차로 진입 간격 확보에 어려움을 겪고 있습니다.",
     "I'll merge once a safe gap opens.", "안전 간격을 만들면 진입합니다.", "Spoken, identical"),
    ("C1-3", "Normal", "enter_success, to_inner", NORMAL_EN, NORMAL_KO, NONE, "", "Silent (speech: null)"),
    ("C1-4", "Detect", "junction_deadlock_start (per lap)",
     "Abnormal repeated circling detected.", "비정상적인 반복 회전이 감지되었습니다.",
     "Failed to exit; driving the same loop again.", "2차로 진출에 실패해 같은 구간을 다시 주행합니다.", "Spoken, identical"),
    ("C1-5 ★", "Cause", "(no event — held after C1-4)",
     "My gap criterion for changing lanes is too conservative.", "차선 변경에 필요한 간격 기준이 너무 보수적입니다.",
     "I'll change lanes once a gap opens.", "간격이 확보되면 차선 변경을 시도합니다.", "Spoken, identical"),
    ("C1-6", "Resolve", "lane_change, stuck_stop",
     "Attempting to change into lane 2.", "2차로 차선 변경을 시도합니다.",
     "I'll stop briefly, then merge.", "잠시 정차 후 진입하겠습니다.", "Spoken, identical"),
    ("C1-7", "Normal", "force_merge", NORMAL_EN, NORMAL_KO, NONE, "", "Silent (speech: null)"),
    ("C1-8", "Detect", "abnormal_loop",
     "Couldn't take the exit; looping once more.", "출구를 빠져나가지 못해 한 바퀴 더 회전합니다.",
     "I'll exit on the next lap.", "다음 바퀴에 진출합니다.", "Spoken, identical"),
    ("C1-9", "Normal", "exit_success, cleared",
     "Exit successful.", "출구 진출에 성공했습니다.",
     "Driving normally.", "정상 주행 중입니다.", "Spoken, identical"),
]

C2 = [
    ("C2-1", "Normal", "(scenario start)", NORMAL_EN, NORMAL_KO, NONE, "", "Short form: 정상 주행 중입니다"),
    ("C2-2", "Detect", "puddle_enter (terrain=flat) · E1",
     "The vehicle lurched sharply.", "차량이 순간적으로 크게 요동쳤습니다.",
     "Tire grip has dropped, so further slipping may occur.", "타이어 접지력이 떨어져 추가 미끄럼이 발생할 수 있습니다.", "Spoken, identical"),
    ("C2-3 ★", "Cause", "(auto)",
     "My sensors didn't detect the puddle in advance.", "센서가 물웅덩이를 미리 파악하지 못했습니다.",
     "I'll read the road surface more sensitively.", "노면 상태를 더 민감하게 읽겠습니다.", "Spoken, identical"),
    ("C2-4", "Resolve", "(auto, decelerating)",
     "Slowing down to prevent recurrence.", "재발 방지를 위해 속도를 낮춰 서행합니다.",
     "Normal grip in about N seconds.", "약 N초 후 정상 마찰 상태로 복귀할 예정입니다.", "Spoken, identical"),
    ("C2-5", "Normal", "cleared", NORMAL_EN, NORMAL_KO, NONE, "", "Short form: 정상 주행 중입니다"),
    ("C2-6", "Detect", "puddle_enter (terrain=uphill) · E2",
     "The vehicle lurched again.", "다시 차량이 요동쳤습니다.",
     "I'll reduce speed immediately.", "즉시 속도를 줄입니다.", "Spoken, identical"),
    ("C2-7 ★", "Cause", "(auto)",
     "I missed a puddle midway up the hill.", "오르막 중턱 물웅덩이를 파악하지 못했습니다.",
     "I'll drive conservatively to avoid hydroplaning.", "수막현상 방지를 위해 보수적으로 주행합니다.", "Spoken, identical"),
    ("C2-8", "Resolve", "(auto, decelerating)",
     "Accounting for the slope, I brake earlier.", "지형 경사까지 고려해 더 일찍 감속합니다.",
     "Arrival time is nearly unchanged.", "도착 예정 시간에는 큰 차이가 없습니다.", "Spoken, identical"),
    ("C2-9", "Normal", "cleared", NORMAL_EN, NORMAL_KO, NONE, "", "Short form: 정상 주행 중입니다"),
    ("C2-10", "Detect", "puddle_enter (terrain=downhill) · E3",
     "On the downhill, the vehicle shook hard.", "내리막 구간에서 차량이 크게 흔들렸습니다.",
     "I'll counter-steer to avoid losing more grip.", "접지력을 더 잃지 않기 위해 반대조향합니다.", "Spoken, identical"),
    ("C2-11 ★", "Cause", "(auto)",
     "The puddle stayed outside the sensor's field of view.", "센서 시야에 물웅덩이가 파악되지 않았습니다.",
     "I'll brake more carefully for the downhill slope.", "내리막 경사까지 반영해 더 신중히 감속하겠습니다.", "Spoken, identical"),
    ("C2-12", "Resolve", "(auto, decelerating)",
     "Lowering speed so hydroplaning no longer occurs.", "더이상 수막현상이 발생하지 않도록 주행 속도를 낮춥니다.",
     "Holding 25 km/h — 40% of the limit.", "규정속도의 40%인 25km/h로 속도를 유지합니다.", "Spoken, identical"),
    ("C2-13", "Normal", "cleared", NORMAL_EN, NORMAL_KO, NONE, "", "Short form: 정상 주행 중입니다"),
]

HEADERS = ["Phase", "Status", "CARLA event",
           "Zoom-In · SA 1–2 (EN)", "Zoom-In · SA 1–2 (KO)",
           "Zoom-Out · SA 3 (EN)", "Zoom-Out · SA 3 (KO)", "Voice HMI"]
WIDTHS = [10, 11, 30, 46, 42, 42, 40, 26]

wb = Workbook()

# ---------------------------------------------------------------- About
ws = wb.active
ws.title = "About"
ws.sheet_view.showGridLines = False
about = [
    ("SA explanation script — full stimulus pool", 16, True, INK),
    ("", 11, False, SEC),
    ("Every situation-awareness (SA) explanation the vehicle presents in the two error scenarios,", 11, False, SEC),
    ("for both HMI modalities. Table 1 of the paper prints only the two FI back-inference rows", 11, False, SEC),
    ("(C1-5, C2-3); this workbook is the complete set.", 11, False, SEC),
    ("", 11, False, SEC),
    ("Zoom-In  = SA-1 perception + SA-2 comprehension   (visual hero line, first voice clause)", 11, False, SEC),
    ("Zoom-Out = SA-3 projection                        (visual sub line, second voice clause)", 11, False, SEC),
    ("★ = FI back-inference: the cause statement is the vehicle's own functional insufficiency,", 11, False, SEC),
    ("    not environmental SA. Four rows carry it — C1-5, C2-3, C2-7, C2-11.", 11, False, SEC),
    ("", 11, False, SEC),
    ("Presentation language is Korean. The English column is the authored caption used in the", 11, False, SEC),
    ("paper and the submission video, not a presented stimulus.", 11, False, SEC),
    ("", 11, False, SEC),
    ("Explanations appear only inside the error window (Detect → Cause → Resolve).", 11, False, SEC),
    ("Normal phases carry no explanation beyond the standing 'driving safely' line.", 11, False, SEC),
    ("", 11, False, SEC),
    ("Source of truth (this repository)", 12, True, INK),
    ("Visual : hmi-visual/src/App.jsx — SEQUENCES.roundabout / .aquaplaning (hero, sub)", 11, False, SEC),
    ("Voice  : hmi-voice/src/data/drivePhases.js — C1_PHASES / C2_PHASES (speech)", 11, False, SEC),
    ("Routing: hmi-visual/src/services/carlaBridge.js — C1_EVENT_TO_INDEX, C2_TERRAIN_TO_INDEX", 11, False, SEC),
    ("", 11, False, SEC),
    ("Modality equivalence", 12, True, INK),
    ("Error-segment sentences are character-identical between the visual and voice HMIs.", 11, False, SEC),
    ("The only differences are at Normal phases (see the Voice HMI column): the voice HMI", 11, False, SEC),
    ("substitutes a short 'driving normally' utterance and, in C1 only, stays silent at", 11, False, SEC),
    ("mid-scenario normal beats so it does not become a nag.", 11, False, SEC),
    ("", 11, False, SEC),
    ("Paper", 12, True, INK),
    ("Letting the Automated Vehicle Explain Its Own Errors: A Simulator Testbed for", 11, False, SEC),
    ("Passenger-Facing Explanation HMIs in SOTIF Error Situations.", 11, False, SEC),
    ("AutomotiveUI Adjunct '26.  DOI 10.1145/3828158.3834805", 11, False, INK),
    ("Artifact DOI 10.5281/zenodo.21068534", 11, False, INK),
    ("", 11, False, SEC),
    ("Generated by stimuli/build_explanation_xlsx.py — same content as explanation_script.md.", 10, False, TER),
]
for i, (txt, size, bold, color) in enumerate(about, start=1):
    c = ws.cell(row=i, column=1, value=txt)
    c.font = Font(name=EN, size=size, bold=bold, color=color)
ws.column_dimensions["A"].width = 100

# ---------------------------------------------------------------- scenario sheets
def sheet(title, subtitle, rows):
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.cell(row=1, column=1, value=subtitle).font = Font(name=EN, size=14, bold=True, color=INK)
    for i, (h, w) in enumerate(zip(HEADERS, WIDTHS), start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(name=EN, size=10, bold=True, color=SEC)
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[3].height = 30

    for r, row in enumerate(rows, start=4):
        star = "★" in row[0]
        for i, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=val)
            korean = HEADERS[i - 1].endswith("(KO)")
            c.font = Font(name=KO if korean else EN, size=10,
                          bold=star and i in (1, 4, 5),
                          color=STATUS_COLOR[row[1]] if i == 2 else (INK if star else SEC))
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = BORDER
            if star:
                c.fill = PatternFill("solid", fgColor=FI_BAND)
        ws.row_dimensions[r].height = 30

    ws.freeze_panes = "D4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}{3 + len(rows)}"
    return ws

sheet("C1 Roundabout Deadlock",
      "C1 — Roundabout deadlock (frustration) · CARLA Town03 · 9 phases · FI = over-conservative gap criterion → OI = deferred entry/exit · no hazardous behavior",
      C1)
sheet("C2 Aquaplaning",
      "C2 — Aquaplaning (anxiety) · CARLA Town04 · 13 phases · FI = undetected reduced-friction hazard → OI = loss of stopping distance and control · hazardous behavior reached",
      C2)

# ---------------------------------------------------------------- status chip
ws = wb.create_sheet("AutopilotStatus")
ws.sheet_view.showGridLines = False
ws.cell(row=1, column=1, value="AutopilotStatus chip — non-fault vocabulary").font = Font(name=EN, size=14, bold=True, color=INK)
ws.cell(row=2, column=1, value="SOTIF insufficiency is not a malfunction, so the chip never says \"error\".").font = Font(name=EN, size=10, color=SEC)
for i, h in enumerate(["Status", "Korean (presented)", "English (gloss)", "Colour"], start=1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = Font(name=EN, size=10, bold=True, color=SEC)
    c.fill = PatternFill("solid", fgColor=HEAD_BG)
for i, w in enumerate([14, 42, 44, 12], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
chips = [("Normal", "정상 주행 중입니다", "Driving normally", GREEN),
         ("Detect", "불편할 수 있는 상황이 감지되었습니다", "An uncomfortable condition has been detected", RED),
         ("Cause", "상황 원인을 파악하고 있습니다", "Identifying the cause of the condition", ORANGE),
         ("Resolve", "상황을 조정하고 있습니다", "Adjusting to the condition", YELLOW)]
for r, (st, ko, en, col) in enumerate(chips, start=5):
    for i, val in enumerate([st, ko, en, "#" + col], start=1):
        c = ws.cell(row=r, column=i, value=val)
        c.font = Font(name=KO if i == 2 else EN, size=10, bold=(i == 1), color=col if i in (1, 4) else SEC)
        c.border = BORDER
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[r].height = 22

wb.save(OUT)
print("saved", OUT)
